#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import asyncio
import argparse
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Set, Tuple, Any, Optional
from dataclasses import dataclass
import time
import json
from tqdm.asyncio import tqdm
import re
from concurrent.futures import ThreadPoolExecutor

# Import our translator module
from translator import AsyncTranslator, TranslationResult, setup_logging

# Define RTL languages
RTL_LANGUAGES = {
    'ar': 'Arabic',
    'fa': 'Persian',
    'he': 'Hebrew',
    'ur': 'Urdu',
    'yi': 'Yiddish',
    'dv': 'Dhivehi',
    'ku': 'Kurdish',
    'ps': 'Pashto',
    'sd': 'Sindhi',
    'ug': 'Uyghur'
}

@dataclass
class TranslationTask:
    """Data class to track translation tasks for Excel files."""
    sheet_name: str
    cell_address: str
    original_text: str
    column_index: int
    row_index: int
    cell_style: dict
    cell_format: Optional[str] = None

@dataclass
class TranslationConfig:
    """Configuration for translation tasks."""
    source_lang: str = 'auto'
    target_lang: str = 'en'
    columns_to_translate: List[int] = None  # 0-based indices
    sheets_to_translate: List[str] = None   # Sheet names
    skip_rows: int = 0                      # Header rows to skip
    preserve_formatting: bool = True
    max_concurrent_requests: int = 10
    batch_size: int = 50
    add_translated_columns: bool = True     # Add new columns instead of replacing
    column_prefix: str = "Translated_"      # Prefix for new columns
    timestamp_output: bool = True           # Add timestamp to output filename
    cache_file: Optional[str] = 'translation_cache.json'
    debug: bool = False

class XlsxTranslator:
    """
    Translate XLSX files content with advanced features:
    - Multiple sheet support
    - Selective column translation
    - Format preservation
    - Progress tracking
    - RTL/LTR language support
    - Translation caching
    - Error handling and reporting
    """
    
    def __init__(self, config: TranslationConfig, logger: Optional[logging.Logger] = None):
        """Initialize the XLSX translator with configuration."""
        self.config = config
        self.logger = logger or setup_logging(
            log_file='xlsx_translation.log', 
            log_level=logging.DEBUG if config.debug else logging.INFO
        )
        
        # Set up the translator
        self.translator = AsyncTranslator(
            max_concurrent_requests=config.max_concurrent_requests,
            logger=self.logger
        )
        
        # Initialize translation metrics
        self.total_cells = 0
        self.translated_cells = 0
        self.errors = 0
        self.skipped_cells = 0
        self.start_time = None
        
        # Load translation cache if available
        self.translation_cache = {}
        if config.cache_file and os.path.exists(config.cache_file):
            try:
                with open(config.cache_file, 'r', encoding='utf-8') as f:
                    self.translation_cache = json.load(f)
                self.logger.info(f"Loaded {len(self.translation_cache)} entries from cache")
            except Exception as e:
                self.logger.warning(f"Failed to load cache file: {str(e)}")
    
    async def translate_xlsx(self, input_file: str, output_file: Optional[str] = None) -> str:
        """
        Translate an XLSX file while preserving structure and formatting.
        
        Args:
            input_file: Path to input XLSX file
            output_file: Path to output XLSX file (optional)
            
        Returns:
            Path to the translated file
        """
        # Auto-generate output filename if not provided
        if not output_file:
            base_name, ext = os.path.splitext(input_file)
            if self.config.timestamp_output:
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                output_file = f"{base_name}_{self.config.target_lang}_{timestamp}{ext}"
            else:
                output_file = f"{base_name}_{self.config.target_lang}{ext}"
        
        self.logger.info(f"Translating file: {input_file} -> {output_file}")
        self.logger.info(f"Configuration: src={self.config.source_lang}, dest={self.config.target_lang}")
        
        self.start_time = time.time()
        
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(input_file)
            sheet_names = workbook.sheetnames
            
            # Filter sheets if specified
            if self.config.sheets_to_translate:
                sheet_names = [name for name in sheet_names if name in self.config.sheets_to_translate]
                
            self.logger.info(f"Processing {len(sheet_names)} sheets: {', '.join(sheet_names)}")
            
            # Collect all translation tasks first
            all_tasks = []
            for sheet_name in sheet_names:
                sheet_tasks = self._collect_translation_tasks(workbook[sheet_name])
                all_tasks.extend(sheet_tasks)
                
            self.total_cells = len(all_tasks)
            self.logger.info(f"Found {self.total_cells} cells to translate")
            
            # Process translations in batches
            translated_tasks = await self._process_translation_tasks(all_tasks)
            
            # Apply translations back to the workbook
            self._apply_translations(workbook, translated_tasks)
            
            # Save the workbook
            workbook.save(output_file)
            
            # Save updated cache
            if self.config.cache_file:
                try:
                    with open(self.config.cache_file, 'w', encoding='utf-8') as f:
                        json.dump(self.translation_cache, f, ensure_ascii=False, indent=2)
                    self.logger.info(f"Updated cache with {len(self.translation_cache)} entries")
                except Exception as e:
                    self.logger.warning(f"Failed to save cache file: {str(e)}")
            
            # Print summary
            elapsed_time = time.time() - self.start_time
            self.logger.info(
                f"Translation completed in {elapsed_time:.2f}s. "
                f"Cells: {self.translated_cells}/{self.total_cells} translated, "
                f"{self.skipped_cells} skipped, {self.errors} errors. "
                f"Success rate: {(self.translated_cells/self.total_cells)*100:.1f}%"
            )
            
            return output_file
            
        except Exception as e:
            self.logger.error(f"Error translating file: {str(e)}", exc_info=True)
            raise
    
    def _collect_translation_tasks(self, sheet) -> List[TranslationTask]:
        """
        Collect cells that need translation from a worksheet.
        
        Args:
            sheet: openpyxl worksheet
            
        Returns:
            List of TranslationTask objects
        """
        tasks = []
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Determine which columns to translate
        columns_to_process = []
        if self.config.columns_to_translate:
            # Convert 0-based indices to 1-based for openpyxl
            columns_to_process = [col + 1 for col in self.config.columns_to_translate 
                                  if col + 1 <= max_col]
        else:
            columns_to_process = list(range(1, max_col + 1))
        
        self.logger.info(f"Sheet '{sheet.title}': Processing {len(columns_to_process)} columns out of {max_col}")
        
        # Process cells row by row
        start_row = self.config.skip_rows + 1  # Convert to 1-based indexing
        
        for row in range(start_row, max_row + 1):
            for col in columns_to_process:
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value
                
                # Skip empty cells, numbers, and formulas
                if (cell_value is None or 
                    isinstance(cell_value, (int, float)) or 
                    (isinstance(cell_value, str) and cell_value.startswith('='))):
                    continue
                    
                # Convert to string if not already
                cell_value = str(cell_value).strip()
                
                # Skip if empty after stripping
                if not cell_value:
                    continue
                
                # Capture formatting info if needed
                cell_style = None
                if self.config.preserve_formatting:
                    cell_style = {
                        'alignment': cell.alignment,
                        'font': cell.font,
                        'fill': cell.fill,
                        'border': cell.border,
                        'number_format': cell.number_format
                    }
                
                # Cell address for logging
                cell_address = f"{get_column_letter(col)}{row}"
                
                tasks.append(TranslationTask(
                    sheet_name=sheet.title,
                    cell_address=cell_address,
                    original_text=cell_value,
                    column_index=col,
                    row_index=row,
                    cell_style=cell_style,
                    cell_format=cell.number_format if hasattr(cell, 'number_format') else None
                ))
        
        self.logger.info(f"Sheet '{sheet.title}': Collected {len(tasks)} cells for translation")
        return tasks
        
    async def _process_translation_tasks(self, tasks: List[TranslationTask]) -> List[Tuple[TranslationTask, TranslationResult]]:
        """
        Process translation tasks in batches.
        
        Args:
            tasks: List of TranslationTask objects
            
        Returns:
            List of (TranslationTask, TranslationResult) tuples
        """
        results = []
        
        # Extract texts and check cache
        texts_to_translate = []
        cached_results = []
        uncached_tasks = []
        uncached_indices = []
        
        for i, task in enumerate(tasks):
            cache_key = f"{task.original_text}:{self.config.source_lang}:{self.config.target_lang}"
            
            if cache_key in self.translation_cache:
                # Use cached translation
                self.logger.debug(f"Cache hit for cell {task.cell_address} in sheet '{task.sheet_name}'")
                cached_result = TranslationResult(
                    original_text=task.original_text,
                    translated_text=self.translation_cache[cache_key],
                    source_language=self.config.source_lang,
                    target_language=self.config.target_lang,
                    success=True
                )
                cached_results.append((task, cached_result))
                self.translated_cells += 1
            else:
                # Need to translate
                texts_to_translate.append(task.original_text)
                uncached_tasks.append(task)
                uncached_indices.append(i)
        
        self.logger.info(f"Found {len(cached_results)} cached translations, {len(uncached_tasks)} to translate")
        
        # Add cached results
        results.extend(cached_results)
        
        # If there are uncached tasks, translate them in batches
        if uncached_tasks:
            batch_size = self.config.batch_size
            total_batches = (len(texts_to_translate) + batch_size - 1) // batch_size
            
            self.logger.info(f"Translating {len(texts_to_translate)} texts in {total_batches} batches")
            
            for i in range(0, len(texts_to_translate), batch_size):
                batch_texts = texts_to_translate[i:i + batch_size]
                batch_tasks = uncached_tasks[i:i + batch_size]
                
                batch_num = i // batch_size + 1
                self.logger.info(f"Processing batch {batch_num}/{total_batches} ({len(batch_texts)} texts)")
                
                # Translate batch
                translations = await self.translator.batch_translate(
                    batch_texts, 
                    src_lang=self.config.source_lang, 
                    dest_lang=self.config.target_lang
                )
                
                # Process results
                for task, translation in zip(batch_tasks, translations):
                    if translation.success and translation.translated_text:
                        # Update cache
                        cache_key = f"{task.original_text}:{self.config.source_lang}:{self.config.target_lang}"
                        self.translation_cache[cache_key] = translation.translated_text
                        self.translated_cells += 1
                    else:
                        self.errors += 1
                        self.logger.warning(
                            f"Failed to translate cell {task.cell_address} in sheet '{task.sheet_name}': "
                            f"{translation.error_message or 'Unknown error'}"
                        )
                    
                    results.append((task, translation))
                
                # Progress update
                progress = (len(results) - len(cached_results)) / len(uncached_tasks) * 100
                elapsed = time.time() - self.start_time
                self.logger.info(
                    f"Batch {batch_num}/{total_batches} completed. "
                    f"Progress: {progress:.1f}%. Time elapsed: {elapsed:.1f}s"
                )
        
        return results
    
    def _apply_translations(self, workbook, translated_tasks: List[Tuple[TranslationTask, TranslationResult]]):
        """
        Apply translations back to the workbook.
        
        Args:
            workbook: openpyxl workbook
            translated_tasks: List of (TranslationTask, TranslationResult) tuples
        """
        # Group tasks by sheet
        tasks_by_sheet = {}
        for task, result in translated_tasks:
            if task.sheet_name not in tasks_by_sheet:
                tasks_by_sheet[task.sheet_name] = []
            tasks_by_sheet[task.sheet_name].append((task, result))
        
        # Process each sheet
        for sheet_name, sheet_tasks in tasks_by_sheet.items():
            sheet = workbook[sheet_name]
            
            # Determine if we need to add new columns
            new_columns = {}
            if self.config.add_translated_columns:
                # Find which original columns need translated counterparts
                original_columns = set(task.column_index for task, _ in sheet_tasks)
                
                # Find the max column in use
                max_col = sheet.max_column
                
                # Create mapping of original column -> translated column
                col_offset = 0
                for original_col in sorted(original_columns):
                    # Add translated column right after the original
                    translated_col = max_col + col_offset + 1
                    new_columns[original_col] = translated_col
                    col_offset += 1
                    
                    # Get column letters for header creation
                    orig_col_letter = get_column_letter(original_col)
                    trans_col_letter = get_column_letter(translated_col)
                    
                    # Copy column width
                    if sheet.column_dimensions[orig_col_letter].width:
                        sheet.column_dimensions[trans_col_letter].width = sheet.column_dimensions[orig_col_letter].width
                    
                    # Add header
                    if self.config.skip_rows > 0:
                        for header_row in range(1, self.config.skip_rows + 1):
                            # Copy header content with translation indicator
                            orig_header_cell = sheet.cell(row=header_row, column=original_col)
                            new_header_cell = sheet.cell(row=header_row, column=translated_col)
                            
                            if header_row == self.config.skip_rows:  # Last header row
                                if orig_header_cell.value:
                                    new_header_cell.value = f"{self.config.column_prefix}{orig_header_cell.value}"
                                else:
                                    new_header_cell.value = f"{self.config.column_prefix}Col {orig_col_letter}"
                            else:
                                new_header_cell.value = orig_header_cell.value
                            
                            # Copy formatting
                            if self.config.preserve_formatting and hasattr(orig_header_cell, 'font'):
                                new_header_cell.font = copy_font_style(orig_header_cell.font)
                                new_header_cell.alignment = copy_alignment(orig_header_cell.alignment)
                                new_header_cell.fill = copy_fill(orig_header_cell.fill)
                                new_header_cell.border = copy_border(orig_header_cell.border)
            
            # Apply translations
            is_rtl = self.config.target_lang in RTL_LANGUAGES
            
            for task, result in sheet_tasks:
                if result.success and result.translated_text:
                    # Determine target cell
                    if self.config.add_translated_columns:
                        col_idx = new_columns[task.column_index]
                    else:
                        col_idx = task.column_index
                    
                    target_cell = sheet.cell(row=task.row_index, column=col_idx)
                    
                    # Apply translation
                    target_cell.value = clean_translation(result.translated_text)
                    
                    # Apply RTL text direction if needed
                    if is_rtl:
                        if not target_cell.alignment:
                            target_cell.alignment = Alignment()
                        target_cell.alignment = Alignment(
                            horizontal='right', 
                            vertical=target_cell.alignment.vertical,
                            textRotation=target_cell.alignment.textRotation,
                            wrapText=True,
                            readingOrder=2  # RTL reading order
                        )
                    
                    # Apply original formatting if needed
                    if self.config.preserve_formatting and task.cell_style:
                        # Skip alignment if it's an RTL language (we've already set it)
                        if not is_rtl:
                            target_cell.alignment = task.cell_style['alignment']
                        target_cell.font = task.cell_style['font']
                        target_cell.fill = task.cell_style['fill']
                        target_cell.border = task.cell_style['border']
                        if task.cell_format:
                            target_cell.number_format = task.cell_format
                else:
                    # Skip failed translations
                    self.skipped_cells += 1

def copy_font_style(font):
    """Copy font style to new cell."""
    return Font(
        name=font.name,
        size=font.size,
        bold=font.bold,
        italic=font.italic,
        color=font.color
    )

def copy_alignment(alignment):
    """Copy alignment to new cell."""
    return Alignment(
        horizontal=alignment.horizontal,
        vertical=alignment.vertical,
        textRotation=alignment.textRotation,
        wrapText=alignment.wrapText,
        shrinkToFit=alignment.shrinkToFit,
        indent=alignment.indent,
    )

def copy_fill(fill):
    """Copy fill to new cell."""
    if fill.patternType:
        return PatternFill(
            patternType=fill.patternType,
            fgColor=fill.fgColor,
            bgColor=fill.bgColor
        )
    return fill

def copy_border(border):
    """Copy border to new cell."""
    return border

def clean_translation(text: str) -> str:
    """Clean up the translated text for Excel."""
    # Remove excessive whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Fix common issues with translations
    text = text.replace('\u200b', '')  # Remove zero-width spaces
    
    # Fix line breaks
    text = text.replace('\\n', '\n')
    
    return text

def detect_excel_type(file_path: str) -> str:
    """Detect if the Excel file is XLSX or XLS format."""
    _, ext = os.path.splitext(file_path)
    return ext.lower()

async def process_file(args):
    """Process a single file with the given arguments."""
    config = TranslationConfig(
        source_lang=args.source_lang,
        target_lang=args.target_lang,
        columns_to_translate=args.columns,
        sheets_to_translate=args.sheets,
        skip_rows=args.skip_rows,
        preserve_formatting=args.preserve_formatting,
        max_concurrent_requests=args.max_requests,
        batch_size=args.batch_size,
        add_translated_columns=not args.replace,
        column_prefix=args.prefix,
        timestamp_output=args.timestamp,
        cache_file=args.cache_file,
        debug=args.debug
    )
    
    translator = XlsxTranslator(config)
    return await translator.translate_xlsx(args.input_file, args.output_file)

async def batch_process(args):
    """Process multiple files in a directory."""
    input_dir = args.input_dir
    output_dir = args.output_dir or input_dir
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Get all Excel files in the input directory
    excel_files = []
    for file in os.listdir(input_dir):
        if file.endswith(('.xlsx', '.xls')):
            excel_files.append(os.path.join(input_dir, file))
    
    logger = setup_logging(log_file='batch_translation.log')
    logger.info(f"Found {len(excel_files)} Excel files to process")
    
    # Process each file
    results = []
    for file_path in excel_files:
        try:
            file_name = os.path.basename(file_path)
            output_path = os.path.join(output_dir, file_name.replace('.', f'_{args.target_lang}.'))
            
            # Create args for this file
            file_args = argparse.Namespace(
                input_file=file_path,
                output_file=output_path,
                source_lang=args.source_lang,
                target_lang=args.target_lang,
                columns=args.columns,
                sheets=args.sheets,
                skip_rows=args.skip_rows,
                preserve_formatting=args.preserve_formatting,
                max_requests=args.max_requests,
                batch_size=args.batch_size,
                replace=args.replace,
                prefix=args.prefix,
                timestamp=args.timestamp,
                cache_file=args.cache_file,
                debug=args.debug
            )
            
            logger.info(f"Processing file: {file_name}")
            output_file = await process_file(file_args)
            results.append((file_path, output_file, True))
            logger.info(f"Successfully translated: {file_name} -> {os.path.basename(output_file)}")
            
        except Exception as e:
            logger.error(f"Error processing {file_path}: {str(e)}")
            results.append((file_path, None, False))
    
    # Print summary
    logger.info("\nBatch Processing Summary:")
    success_count = sum(1 for _, _, success in results if success)
    logger.info(f"Total files: {len(results)}")
    logger.info(f"Successfully translated: {success_count}")
    logger.info(f"Failed: {len(results) - success_count}")
    
    return results

def main():
    """Main entry point for the translator."""
    parser = argparse.ArgumentParser(description='Excel File Translator')
    
    # Input/output options
    parser.add_argument('input_file', help='Input Excel file path', nargs='?')
    parser.add_argument('-o', '--output-file', help='Output Excel file path')
    parser.add_argument('-d', '--input-dir', help='Input directory for batch processing')
    parser.add_argument('-od', '--output-dir', help='Output directory for batch processing')
    
    # Translation options
    parser.add_argument('-s', '--source-lang', default='auto', help='Source language code (default: auto)')
    parser.add_argument('-t', '--target-lang', default='en', help='Target language code')
    parser.add_argument('-c', '--columns', type=int, nargs='+', help='Columns to translate (0-based indices)')
    parser.add_argument('-S', '--sheets', nargs='+', help='Sheets to translate (names)')
    parser.add_argument('-r', '--skip-rows', type=int, default=1, help='Number of header rows to skip (default: 1)')
    
    # Formatting options
    parser.add_argument('--no-preserve-formatting', dest='preserve_formatting', action='store_false',
                        help='Do not preserve cell formatting')
    parser.add_argument('--replace', action='store_true', help='Replace original content instead of adding columns')
    parser.add_argument('--prefix', default='Translated_', help='Prefix for translated column headers')
    parser.add_argument('--no-timestamp', dest='timestamp', action='store_false',
                        help='Do not add timestamp to output filename')
    
    # Performance options
    parser.add_argument('-m', '--max-requests', type=int, default=10, 
                        help='Maximum concurrent translation requests (default: 10)')
    parser.add_argument('-b', '--batch-size', type=int, default=50,
                        help='Translation batch size (default: 50)')
    parser.add_argument('--cache-file', default='translation_cache.json',
                        help='Translation cache file (default: translation_cache.json)')
    parser.add_argument('--no-cache', dest='cache_file', action='store_const', const=None,
                        help='Disable translation caching')
    
    # Debug options
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Validate arguments
    if args.input_dir:
        if not os.path.isdir(args.input_dir):
            print(f"Error: Input directory '{args.input_dir}' does not exist")
            return 1
        asyncio.run(batch_process(args))
    elif args.input_file:
        if not os.path.isfile(args.input_file):
            print(f"Error: Input file '{args.input_file}' does not exist")
            return 1
        asyncio.run(process_file(args))
    else:
        parser.print_help()
        return 1
    
    return 0

if __name__ == "__main__":
    # Set up asyncio policies for Windows if needed
    if hasattr(asyncio, 'WindowsSelectorEventLoopPolicy') and hasattr(asyncio, 'set_event_loop_policy'):
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
    exit(main())
